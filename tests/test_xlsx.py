import os
from openpyxl import load_workbook

from conftest import PROJECT_RESOURCE_PATH


# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_xlsx_processing():
    xlsx_file = os.path.join(PROJECT_RESOURCE_PATH, 'file_example_XLSX_50.xlsx')
    workbook = load_workbook(xlsx_file)
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)
    assert sheet.cell(row=3, column=2).value == 'Mara'
    assert sheet.cell(row=51, column=8).value == 6125
